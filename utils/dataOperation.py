import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from collections import Counter

class DataHandle:
    def __init__(self, save_path, file_name):
        self.save_path = save_path
        self.file_name = file_name
        self.chart = {
            "Choice": ["机票代码", "选择时间(s)", "对比机票代号", "测试时间"],
            "Finder": ["机票代号", "题目", "选择时间(s)", "正确情况", "测试时间"]
        }

    def mk_data_set(self):
        # 检测文件是否存在
        try:
            workbook = openpyxl.load_workbook(filename=f"{self.save_path}\{self.file_name}")
        except FileNotFoundError:
            # 文件不存在时创建文件并添加表格
            workbook = Workbook()

            if "Sheet" in workbook.sheetnames:
                workbook.remove(workbook["Sheet"])

            for sheet_name, headers in self.chart.items():
                sheet = workbook.create_sheet(title=sheet_name)
                for col_num, header in enumerate(headers, 1):
                    col_letter = get_column_letter(col_num)
                    sheet[f"{col_letter}1"] = header
            workbook.save(filename=f"{self.save_path}\{self.file_name}")

    def add_data(self, sheet_name, data):
        workbook = openpyxl.load_workbook(filename=f"{self.save_path}\{self.file_name}")
        sheet = workbook[sheet_name]
        max_row = sheet.max_row + 1

        for col_num, header in enumerate(self.chart[sheet_name], 1):
            col_letter = get_column_letter(col_num)
            sheet[f"{col_letter}{max_row}"] = data.get(header, "")

        workbook.save(filename=f"{self.save_path}\{self.file_name}")

    def read_data(self, sheet_name):
        # 这个传表单索引号
        workbook = openpyxl.load_workbook(filename=f"{self.save_path}/{self.file_name}")

        sheet_name = list(self.chart.keys())[sheet_name -1]
        # 检查表是否存在
        if sheet_name not in workbook.sheetnames:
            # 表不存在，返回空数据
            return self.chart[sheet_name], []

        sheet = workbook[sheet_name]
        headers = self.chart[sheet_name]
        data = []

        for row in sheet.iter_rows(min_row=2, max_row=sheet.max_row, values_only=True):
            # 将每行数据以元组形式添加到data列表
            data.append(tuple(row))

        return headers, data

    def del_data(self, sheet_name):
            workbook = openpyxl.load_workbook(filename=f"{self.save_path}/{self.file_name}")
            
            sheet_name = list(self.chart.keys())[sheet_name - 1]

            if sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet.delete_rows(2, sheet.max_row)
                workbook.save(filename=f"{self.save_path}/{self.file_name}")
    
    def counter_data(self, sheet_name):
        headers, data = self.read_data(sheet_name)
        result = {header: Counter() for header in headers if header != "测试时间"}

        for row in data:
            for col_num, value in enumerate(row, 1):
                header = headers[col_num - 1]
                if header == "正确情况":
                    # If the column is "正确情况", count occurrences of True
                    if value:
                        result["正确情况"][row[0]] += 1
                elif header == "选择时间(s)":
                    # If the column is "选择时间(s)", sum the times for each "机票代号"
                    result["选择时间(s)"][row[0]] += float(value) if value else 0.0
                elif header != "测试时间" and header != "机票代码":
                    # For other columns, count occurrences as usual
                    result[header][value] += 1

        # Convert the "选择时间(s)" Counter to a dictionary with the sum for each "机票代号"
        result["选择时间(s)"] = dict(result["选择时间(s)"])

        return result